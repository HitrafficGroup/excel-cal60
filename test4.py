from openpyxl import load_workbook
import re
nombre_archivo_existente = '621SISDATCalidadServicio ABR_2022.xlsx'
libro_existente = load_workbook(nombre_archivo_existente)
# cal_stadist = []
# workbook = load_workbook('pruebas2.xlsx',data_only=True)
# nombre_hoja = "Calidad de Servicio Técnico"
# sheet_target = workbook.active
# parametros_sisdat = {'ttik':0,'fmik':0}
# parametros_sisdat['fmik'] = sheet_target['I10'].value
# parametros_sisdat['ttik'] = sheet_target['J10'].value
# for i in range(17,330):
#     aux_cell = sheet_target[f'C{i}'].value
#     aux_fmi = sheet_target[f'I{i}'].value
#     aux_ttk = sheet_target[f'J{i}'].value
#     toal_fmik = 0
#     total_ttik = 0
#     empty_dict = {}
#     if aux_cell != None:
#         name_descompuesto  = re.search(r'\((.*?)\)', aux_cell).group(1)
#         empty_dict['name'] = name_descompuesto
#         empty_dict['fmik'] = aux_fmi
#         empty_dict['ttik'] = aux_ttk
#         cal_stadist.append(empty_dict)
#     else:
#         break
# print(cal_stadist)
