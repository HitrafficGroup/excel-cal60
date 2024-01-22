# File: main.py
import sys
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import QApplication,QFileDialog,QMessageBox,QVBoxLayout,QFrame,QLabel,QWidget
from PySide6.QtCore import QFile, QIODevice,Qt
from PySide6.QtGui import QImage,QPixmap
import os
from openpyxl import load_workbook
import xlrd
import re
import matplotlib as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure
import numpy as np
import openpyxl
from openpyxl.chart import BarChart, Reference
#some init vars
path_file1 = ''
path_file2 = ''
path_file3 = ''

db_wb = load_workbook('db_alimentadores.xlsx',data_only=True)
sheet_db = db_wb.active
lista_names = []
for i in range(6,330):
    aux_name = sheet_db[f'B{i}'].value
    if aux_name == None:
        break
    else:
        name_descompuesto  = re.search(r'\((.*?)\)', aux_name).group(1)
        lista_names.append(name_descompuesto)


all_datos = []
current_target = {}
parametros_sisdat = {'ttik':0,'fmik':0}

class MatplotlibWidget(QWidget):
    def __init__(self, parent=None):
        super(MatplotlibWidget, self).__init__(parent)

        self.initUI()

    def initUI(self):
        # Crear una figura de Matplotlib
        self.fig = Figure(figsize=(5, 4), dpi=100)
        self.canvas = FigureCanvas(self.fig)
        # Crear datos para el gráfico
        x = np.linspace(0, 5, 100)
        y = np.sin(x)
        # Dibujar en la figura
        ax = self.fig.add_subplot(111)
        ax.plot(x, y)
        ax.set_title('Gráfico de Matplotlib')
        layout = QVBoxLayout(self)
        layout.addWidget(self.canvas)

def excelGenerar():
    global current_target
    global all_datos

    workbook = openpyxl.Workbook()
    ws = workbook.active
    format_data = [
            ('','CAL-60','','SISDAT','','ADMS',''),('Alimentador','FMIK', 'TTIK','FMIK','TTIK','FMIK', 'TTIK')
                    ]
    for data in all_datos:
        new_dato = (data['name'],data['file1'][0],data['file1'][1],data['file2'][0],data['file2'][1],data['file3'][0],data['file3'][1])
        format_data.append(new_dato)


    rows = format_data


    for row in rows:
        ws.append(row)


    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')


    counter = 1
    for i in range(3,len(rows)+1):
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = rows[i-1][0]
        chart1.y_axis.title = 'V1.1'
        chart1.x_axis.title = 'Sample length (mm)'
        data = Reference(ws, min_col=2, min_row=i, max_row=i, max_col=7)
        cats = Reference(ws, min_col=1, min_row=i, max_row=i)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.shape = 4
        ws.add_chart(chart1, f"I{counter}")
        counter +=19
    workbook.save("grafico_barras.xlsx")

def auxGenerarExcel():
    workbook = openpyxl.Workbook()
    ws = workbook.active
    format_data = [
                        ('','CAL-60','','SISDAT','','ADMS',''),
                        ('Alimentador','FMIK', 'TTIK','FMIK','TTIK','FMIK', 'TTIK')
                    ]
    for data in all_datos:
        new_dato = (data['name'],data['file1'][0],data['file1'][1],data['file2'][0],data['file2'][1],data['file3'][0],data['file3'][1])
        format_data.append(new_dato)
    rows = format_data
    for row in rows:
        ws.append(row)
    ws.merge_cells('B1:C1')
    ws.merge_cells('D1:E1')
    ws.merge_cells('F1:G1')
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = 'Graficas de Alimentadores'
    chart1.y_axis.title = 'Valores'
    chart1.x_axis.title = f'Alimentadores Totales {len(format_data)}'
    data = Reference(ws, min_col=2, min_row=2, max_row=len(rows), max_col=7)
    cats = Reference(ws, min_col=1, min_row=3, max_row=len(rows))
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    ws.add_chart(chart1, f"I3")
    workbook.save("grafico_barras.xlsx")


def setSrcPath(file):
    global path_file1
    global path_file2
    global path_file3
    fname = QFileDialog.getOpenFileName()
    ruta = fname[0]
    nombre_archivo = os.path.basename(ruta)
    if fname[0][-5:] == ".xlsx" or  fname[0][-4:] == ".xls":
        if file == 1:
            window.txtFile1.setText(nombre_archivo)
            path_file1 = fname[0]
        elif file == 2:
            window.txtFile2.setText(nombre_archivo)
            path_file2 = fname[0]
        else:
            window.txtFile3.setText(nombre_archivo)
            path_file3 = fname[0]
    else:
        alerta = QMessageBox()
        alerta.setWindowTitle('Alerta')
        alerta.setText('Seleccione un Archivo que sea excel')
        alerta.setIcon(QMessageBox.Warning)
        alerta.setStandardButtons(QMessageBox.Ok)
        alerta.setDefaultButton(QMessageBox.Ok)
        alerta.exec()


def processExcel():
    window.listData.clear()
    global path_file1
    global path_file2
    global path_file3
    global all_datos
    global parametros_sisdat
    cal_60_aux = []
    cal_stadist = []
    cal_ttki = []
    if path_file1[-4:] == ".xls" or path_file1[-5:] == ".xlsx":
        libro_trabajo = xlrd.open_workbook(path_file1)
        # Especifica el índice o el nombre de la hoja que deseas utilizar
        nombre_hoja = "Calidad de Servicio Técnico"
        sheet = libro_trabajo.sheet_by_name(nombre_hoja)
        #print('celda 13 y 40',sheet.cell_value(13, 40))
        cal_60 = []
        for i in range(13,500):
            try:
                aux_cell = sheet.cell_value(i, 20)
                fmi = sheet.cell_value(i, 39)
                tki = sheet.cell_value(i, 40)
                empty_dict_cal60 = {}
                if aux_cell != '':
                    name_descompuesto  = re.search(r'\((.*?)\)', aux_cell).group(1)
                    empty_dict_cal60['name'] = name_descompuesto
                    empty_dict_cal60['fmik'] = fmi
                    empty_dict_cal60['ttik'] = tki
                    cal_60.append(empty_dict_cal60)
            except IndexError:
                break
        for i in lista_names:
            new_dict = {'name':i,'ttik':0,'fmik':0}
            counter = 0
            for j in cal_60:
                if j['name'] == i:
                    counter +=1
                    new_dict['fmik'] = new_dict['fmik'] + float(j['fmik']) 
                    new_dict['ttik'] = new_dict['ttik'] + float(j['ttik'])
            if counter != 0:
                cal_60_aux.append(new_dict)
    else:
        pass

    if path_file2[-4:] == ".xls" or path_file2[-5:] == ".xlsx":
    
        workbook = load_workbook(path_file2,data_only=True)
        nombre_hoja = "Calidad de Servicio Técnico"
        sheet_target = workbook.active
        parametros_sisdat['fmik'] = sheet_target['I10'].value
        parametros_sisdat['ttik'] = sheet_target['J10'].value
        for i in range(17,330):
            aux_cell = sheet_target[f'C{i}'].value
            aux_fmi = sheet_target[f'I{i}'].value
            aux_ttk = sheet_target[f'J{i}'].value
            toal_fmik = 0
            total_ttik = 0
            empty_dict = {}
            if aux_cell != None:
                name_descompuesto  = re.search(r'\((.*?)\)', aux_cell).group(1)
                empty_dict['name'] = name_descompuesto
                empty_dict['fmik'] = aux_fmi
                empty_dict['ttik'] = aux_ttk
                cal_stadist.append(empty_dict)
            else:
                break

    else:
        pass
    
    if path_file3[-4:] == ".xls" or path_file3[-5:] == ".xlsx":
        workbook = xlrd.open_workbook(path_file3)
        # Especifica el índice o el nombre de la hoja que deseas utilizar
       

        # Accede a la hoja deseada
        sheet = workbook.sheet_by_index(0)

        for i in range(9,750):
            try:
                aux_cell = sheet.cell_value(i, 7)
                if aux_cell != '':
                    name_descompuesto = aux_cell.split("_")
                    aux_dic = {}
                    if len(name_descompuesto) >=2:
                        for x in range(i,i+50):
                            total_aux = sheet.cell_value(x, 10)
                            if total_aux == 'Total Alimentador':
                                aux_dic['fmik'] = sheet.cell_value(x, 22)
                                aux_dic['ttik'] = sheet.cell_value(x, 24)
                                break
                        aux_dic['name'] = name_descompuesto[2].upper()
    
                        cal_ttki.append(aux_dic)
            except IndexError:
                break

    else:
        pass
    # en esta linea concatanamos todos los datos
    nombres_selectos = []
    for x in lista_names:
        c1 = False
        c2 = False
        c3 = False
        aux_1 = {'name':x,'file1':[0,0],'file2':[0,0],'file3':[0,0]}
        for d1 in cal_60_aux:
            if d1['name'] == x:
                aux_1['file1'] = [d1['fmik'],d1['ttik']]
                c1 = True
                break
        for d2 in cal_stadist:
            if d2['name'] == x:
                aux_1['file2'] = [d2['fmik'],d2['ttik']]
                c2 = True
                break
        for d3 in cal_ttki:
            if d3['name'] == x:
                aux_1['file3'] = [d3['fmik'],d3['ttik']] 
                c3 = True
                break
        if c1 and c2 and c3:
            nombres_selectos.append(aux_1)
            all_datos.append(aux_1)
    lista_ordenada = sorted(nombres_selectos, key=lambda x: x['name'])
    if len(nombres_selectos) > 0:
        for name in lista_ordenada:
            window.listData.addItem(name['name'])
    window.ctd.setText(str(len(nombres_selectos)))
    window.fmik_total.setText(str(parametros_sisdat['fmik']))
    window.ttik_total.setText(str(parametros_sisdat['ttik']))

def nombreSelected():
    global all_datos
    global current_target
    ###
    current_data = {}
    selected_value = window.listData.currentText()
    for i in all_datos:
        if i['name'] == selected_value:
            current_data = i
    current_target = current_data
    window.alimentador.setText(current_data['name'])
    window.cal1.setText(str(current_data['file1'][0]))
    window.cal2.setText(str(current_data['file1'][1]))
    window.sis1.setText(str(current_data['file2'][0]))
    window.sis2.setText(str(current_data['file2'][1]))
    window.adms1.setText(str(current_data['file3'][0]))
    window.adms2.setText(str(current_data['file3'][1]))

    updateGraph(current_data)


def updateGraph(current_data):
        # Obtener el widget Matplotlib y la figura
        categorias = ['cal60 fmik', 'cal60 ttik', 'sisdat fmik', 'sisdat ttik','adms fmik', 'adms ttik']
        valores = [current_data['file1'][0], current_data['file1'][1], current_data['file2'][0], current_data['file2'][1],current_data['file3'][0],current_data['file3'][1]]
        frame = window.graph
        matplotlib_widget = frame.layout().itemAt(0).widget()
        fig = matplotlib_widget.fig
        # Actualizar los datos y redibujar el gráfico
        ax = fig.axes[0]
        ax.clear()
        ax.bar(categorias, valores)
        ax.set_xlabel('Categorías')
        ax.set_ylabel('Valores')
        ax.set_title('Gráfico de Barras')
        fig.canvas.draw()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ui_file_name = "mainwindow.ui"
    ui_file = QFile(ui_file_name)
    if not ui_file.open(QIODevice.ReadOnly):
        print(f"Cannot open {ui_file_name}: {ui_file.errorString()}")
        sys.exit(-1)
    loader = QUiLoader()
    window = loader.load(ui_file)
    window.btnFile1.clicked.connect(lambda: setSrcPath(1))
    window.btnFile2.clicked.connect(lambda: setSrcPath(2))
    window.btnFile3.clicked.connect(lambda: setSrcPath(3))
    window.btnProcess.clicked.connect(lambda: processExcel())
    window.listData.currentIndexChanged.connect(lambda: nombreSelected())
    window.btnExcel.clicked.connect(lambda: auxGenerarExcel())
    matplotlib_widget = MatplotlibWidget(window.graph)
    layout = QVBoxLayout()
    layout.addWidget(matplotlib_widget)
    window.graph.setLayout(layout)

    ui_file.close()
    if not window:
        print(loader.errorString())
        sys.exit(-1)
    window.show()

    sys.exit(app.exec())


    #listado package